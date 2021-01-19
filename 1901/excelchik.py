from openpyxl import Workbook, load_workbook

'''
table = Workbook() <- создать таблицу
sheet = table.active <- выбрать активный лист

sheet['C6'] = 'smth' <- написать что-то в ячейке С6
sheet['B2'] = 'Это Б2' <- написать что-то в ячейке B2

table.save('my_table.xlsx') <- сохранить таблицу
'''
table = load_workbook('my_table.xlsx')
sheet = table.active
print(f'Название листа {sheet.title}')
print(f'Название листов таблицы: {table.sheetnames}')
print(f'В ячейке В2: {sheet["C6"].value}')

for value in sheet.iter_rows(
    min_col=1,
    max_row=5,
    values_only=True
):
    print(f'Cтрока: {value}')